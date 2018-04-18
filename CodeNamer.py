import csv

#単語リストを読み込み
#形容詞リスト作成
adj = open('adjectives.csv','r')
reader_adj = csv.reader(adj)
adj_list = list(reader_adj)
adj.close()

#名詞リスト作成
noun = open('nounlist.csv','r')
reader_noun = csv.reader(noun)
noun_list = list(reader_noun)
noun.close()

import random
from time import sleep
#形容詞からランダムに１つ，名詞からランダムに１つ単語を選び結合・出力
flag = True
while flag:
    adj_index = random.randrange(len(adj_list))
    noun_index = random.randrange(len(adj_list))
    codename = adj_list[adj_index][0]+' '+noun_list[noun_index][0]
    print(codename)
    if input("change the words?(y/n):") == 'n':
        flag = False

#Excelファイルにコードネームと日付を書き込み
import openpyxl as px
from datetime import datetime as dt
wb = px.load_workbook('codename.xlsx')
ws = wb.active
#何も書かれていないセルを見つける
cell_flag = False
i = 0
while cell_flag == False:
    i = i+1
    if ws.cell(row=i, column=1).value == None:
        ws.cell(row=i, column=1).value = codename
        ws.cell(row=i, column=2).value = dt.now()
        cell_flag = True
        
print("---------------------------------------------------------------")
print("Selected codename has been saved successfully")
wb.save('codename.xlsx')
sleep(3)

